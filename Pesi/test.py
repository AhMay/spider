import requests
import json
from requests.cookies import RequestsCookieJar
from Pesi_common import *
from datetime import datetime
import openpyxl

headers = {
        'Accept': 'text/javascript, application/javascript, application/ecmascript, application/x-ecmascript, */*; q=0.01',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Connection': 'keep-alive',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36',
        'X-Requested-With': 'XMLHttpRequest'
    }

def demoExcel():
        wb_name = "永辉全国数据"+ datetime.now().date().strftime("%Y%m")+".xlsx"
        wb = openpyxl.Workbook()
        wb.create_sheet("到货汇总",0)
        wb.create_sheet("到货情况明细",1)
        wb.save(wb_name)
        wb.close()
        wb = openpyxl.load_workbook(wb_name)
        ws = wb['到货汇总']

        for i, header in enumerate(order_sheet_headers.keys()):
                ws.cell(row=1,column=i+1,value=header)

        ws = wb['到货情况明细']

        for i, header in enumerate(order_grid_sheet_headers.keys()):
                ws.cell(row=1, column=i + 1, value=header)
        wb.save(wb_name)

def add_row():
        wb_name = "永辉全国数据" + datetime.now().date().strftime("%Y%m") + ".xlsx"
        wb = openpyxl.load_workbook(wb_name)
        ws = wb['到货汇总']
        ws.append(["1","2","3","4","5","6","7","8","9","10","11","12","13"])
        wb.save(wb_name)




def demoweb():
        from selenium import webdriver
        import os
        import time

        browser = webdriver.Chrome()
        browser.maximize_window()
        browser.get("http://glzx.yonghui.cn/newvssportal/login.html")
        f = input('输入密码后输入任意字符')
        print(browser.current_url)
        cookies = browser.get_cookies()
        signToken = browser.get_cookie('signToken')
        print(signToken)
        print([(c['name'],c['value'].decode('utf-8'))for c in cookies])
        browser.close()
        order_url = "http://glzx.yonghui.cn/newvss-finance-service/vssService/OrderGoodsAction.do?method=searchOrderoodsTotal"
        rand_str = randomWord(False,6)
        random_word = get_sign(appId,rand_str,signToken['value'])
        print(random_word)

        data ={
        "appId": "GLZX_03",
        "random": rand_str,
        "sign":random_word,
        'data': '{"shop":[],"orderDateStart":"2020-01-15","orderDateEnd":"2020-02-14","venderCode":["20002388"],"sheetId":""}'

        }
        pesi_cookieJar = RequestsCookieJar()
        for cookie in cookies:
                pesi_cookieJar.set(cookie['name'],cookie['value'])
        response = requests.post(order_url,headers=headers,data=data,cookies=pesi_cookieJar)
        result = json.loads(response.text)

        order_detail_url = "http://glzx.yonghui.cn/newvss-finance-service/vssService/OrderGoodsAction.do?method=searchOrderGoodsDetail"
        rand_str = randomWord(False,6)
        random_word = get_sign(appId,rand_str,signToken['value'])

        data ={
        "appId": "GLZX_03",
        "random": rand_str,
        "sign":random_word,
        'data': '{"sheetId":"4050054749","source":"0"}'
        }
        response = requests.post(order_detail_url,headers=headers,data=data,cookies=pesi_cookieJar)
        result = json.loads(response.text)

        order_grid_url = "http://glzx.yonghui.cn/newvss-finance-service/vssService/OrderGoodsAction.do?method=orderGoodsDetailGrid"
        rand_str = randomWord(False,6)
        random_word = get_sign(appId,rand_str,signToken['value'])

        data ={
        "appId": "GLZX_03",
        "random": rand_str,
        "sign":random_word,
        'data': '{"sheetId":"4050054749","source":"0"}'
        }
        response = requests.post(order_grid_url,headers=headers,data=data,cookies=pesi_cookieJar)
        result = json.loads(response.text)
        print('here')


demoweb()

def test():
        from selenium import webdriver
        import os
        import time
        try:
                browser = webdriver.Chrome()
                browser.get(logon_url)
                time.sleep(10)
        except Exception as e:
                print()
