import requests
import json
from requests.cookies import RequestsCookieJar
from datetime import datetime
import openpyxl
import os
from selenium import webdriver
import time
from Pesi_common import *
# input starttime
#input endtime
#logon the Pesi
#get the verder list
# get the order data
#get the detail data

#登录首页，获取登录后页面的cookie和signToken
def login():
    browser = webdriver.Chrome()
    browser.maximize_window()
    browser.get(logon_url)
    current_url = browser.current_url
    wait = 50
    while ("login.html" in current_url) and wait >=0:
        time.sleep(5)
        wait = wait -5
        current_url = browser.current_url
    if wait <=0:
        return (None,None)
    cookies = browser.get_cookies()
    browser.close()
    return cookies

def swith_vender(vendercode,cookies):
    browser = webdriver.Chrome()
    browser.maximize_window()
    browser.get(logon_url)
    for cookie in cookies:
        browser.add_cookie({'name':cookie['name'],'value':cookie['value']})
    browser.get(vender_switch_url)
    time.sleep(5)
    count = 10
    while(True):
        adv_element = browser.find_element_by_css_selector("p.user-name")
        if adv_element is not None:
            adv_element.click()
            time.sleep(3)
            vender_element = browser.find_element_by_css_selector("div.change-wrap")
            venders = browser.find_elements_by_css_selector("ul.change-content >li")
            for vender in venders:
                print(vender.text)
                if vender.text.strip().startswith(vendercode):
                    vender.click()
                    break

            break
        else:
            time.sleep(3)
            count  = count-3
            if count <= 0:
                break

    cookies = browser.get_cookies()
    browser.close()
    return cookies

def set_requestData(signToken,datavalue=None):
    rand_str = randomWord(False, 6)
    random_word = get_sign(appId, rand_str, signToken)
    data = {
        "appId": appId,
        "random": rand_str,
        "sign": random_word
    }
    if datavalue is not None:
        data['data'] = str(datavalue)

    return data

def post(url,data,cookies):
    response = requests.post(url, headers=headers, data=data, cookies=cookies)
    result = json.loads(response.text)
    return result

def get_venders(signToken,cookies):
    #获取供应商列表
    cookieJar = RequestsCookieJar()
    for cookie in cookies:
        cookieJar.set(cookie['name'],cookie['value'])
    data = set_requestData(signToken)
    result = post(vender_url,data,cookieJar)
    venders =[]
    for vender in result['data']:
        venders.append(vender['venderCode'])
    return venders

def get_orders(signToken,cookies,orderDateStart,orderDateEnd,venderCode,shop=[],sheetId=""):
    cookieJar = RequestsCookieJar()
    for cookie in cookies:
        cookieJar.set(cookie['name'], cookie['value'])
    qorderdata ={"shop":shop,
                 "orderDateStart":orderDateStart,
                 "orderDateEnd":orderDateEnd,
                 "venderCode":[venderCode,],
                 "sheetId":sheetId}
    data = set_requestData(signToken,qorderdata)
    result = post(order_url,data,cookieJar)
    return result['rows']

def get_details(signToken,cookies,sheetId,source='0'):
    cookieJar = RequestsCookieJar()
    for cookie in cookies:
        cookieJar.set(cookie['name'], cookie['value'])
    qdata ={
        "sheetId":sheetId,
        "source":source
    }
    data = set_requestData(signToken,qdata)
    result = post(order_grid_url,data,cookieJar)
    return result['rows']

def genernate_excel(wb_name=None):
    #生成当月的数据表
    if wb_name is None:
        wb_name = "永辉全国数据" + datetime.now().strftime("%Y%m%d%H%M") + ".xlsx"

    print("新生成的excel文件为", wb_name)
    if os.path.exists(wb_name):
        os.remove(wb_name)
    wb = openpyxl.Workbook()
    wb.create_sheet("到货汇总", 0)
    wb.create_sheet("到货情况明细", 1)
    wb.save(wb_name)
    wb.close()
    wb = openpyxl.load_workbook(wb_name)
    ws = wb['到货汇总']
    for i, header in enumerate(order_sheet_headers.keys()):
        ws.cell(row=1, column=i + 1, value=header)

    ws = wb['到货情况明细']
    for i, header in enumerate(order_grid_sheet_headers.keys()):
        ws.cell(row=1, column=i + 1, value=header)

    wb.save(wb_name)
    wb.close()
    return wb_name

def write_to_sheet(wb_name,sheet_name,objs,referHeaders):
    wb = openpyxl.load_workbook(wb_name)
    ws = wb[sheet_name]
    for obj in objs:
        vals =[]
        for header in referHeaders.keys():
            vals.append(obj[referHeaders[header]])
        ws.append(vals)
    wb.save(wb_name)



if __name__ == '__main__':
   # wb_name = genernate_excel()
   while(True):
        orderDateStart = input("请输入想要查询的开始日期(格式：2020-01-15):")
        if orderDateStart.strip() == "":
            orderDateStart =datetime.now().strftime("%Y-%m-%d")
        orderDateEnd = input("请输入想要查询的结束日期:(格式：2020-01-15):")
        if orderDateEnd.strip() == "":
            orderDateEnd =datetime.now().strftime("%Y-%m-%d")

        print("将要打开登录页面，请稍等输入用户名和密码，并输入验证码登录。")
        cookies = login()
        if cookies is None:
            raise Exception("登录失败！")
        print("登录成功，将要生成Excel文件")
        signToken = [ cookie['value'] for cookie in cookies if cookie['name'] =='signToken'][0]
        venders = get_venders(signToken,cookies)
        wb_name ="永辉" +"_".join(venders)+orderDateStart+"到"+orderDateEnd+".xlsx"
        wb_name = genernate_excel(wb_name)
        current_vender = [ cookie['value'] for cookie in cookies if cookie['name'] =='venderCode'][0]
        venders.remove(current_vender)
        venders.insert(0,current_vender)
        print("拿到供应商代码列表:", venders)
        for vendercode in venders:
            if vendercode != current_vender:
                cookies=swith_vender(vendercode,cookies)
            orders = get_orders(signToken,cookies,orderDateStart,orderDateEnd,vendercode)
            print("得到该供应商：",vendercode,len(orders),"条")
            #todo write orders info to 到货汇总sheet
            if len(orders) != 0:
                write_to_sheet(wb_name,"到货汇总",orders,order_sheet_headers)
            for order in orders:
                details = get_details(signToken,cookies,order['sheetid'])
                print("订单",order["sheetid"],"的供货明细：", len(details), "条")
                #todo write details to 到货情况明细 sheet
                if len(details) !=0:
                    write_to_sheet(wb_name, "到货情况明细", details, order_grid_sheet_headers)

        print("该账户的订单信息已经完成。")
        input("输入任意字符，继续下一个账户的订单内容获取。结束请关闭窗口....")
